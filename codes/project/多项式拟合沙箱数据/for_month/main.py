#coding:utf-8
import pandas as pd
import numpy as np
from openpyxl import Workbook
from functions_sandbox_data import GetAllDaysOfOneMonth, FillnaRank, FillnaNewuser
from functions_sandbox_data import FilterOutliers, FittingRevised

working_space = '/Users/Alas/Documents/TD/iOS_Rank_Estimated_Model_V1/jupyter_sandbox_data/project/'
path_to_write = working_space + 'model/'

# 获取免费排行榜数据
path_top_free_rank = working_space + 'top_free_rank_of_apps_in_sdk_201404-201704.xlsx'
df_rank_all = pd.read_excel(path_top_free_rank, sheetname=0, index_col=1)
df_rank_all.fillna(0, inplace=True)
#print df_rank_all.head(), df_rank_all.tail()

# 保持读取的Excel文件的日期格式与排名数据的时间格式一致
list_new_columns = list(df_rank_all.columns[:3])
for col in df_rank_all.columns[3:]:
    list_new_columns.append(col.strftime('%Y-%m-%d'))
df_rank_all.columns = list_new_columns
#print df_rank_all.head(), df_rank_all.tail()

# 截取2017-01到2017-04的天日期列表
list_days_needed = GetAllDaysOfOneMonth('2017-01') + GetAllDaysOfOneMonth('2017-02') +\
                    GetAllDaysOfOneMonth('2017-03') + GetAllDaysOfOneMonth('2017-04')
df_rank_needed = df_rank_all[list_days_needed]
#print df_rank_needed.head(), df_rank_needed.tail()

wb = Workbook()
ws = wb.active
ws_fillna_rank = wb.create_sheet()
ws_fillna_newuser = wb.create_sheet()
ws.title = 'sort_by_newuser_ios'
ws_fillna_rank.title = 'fillna_and_sort_by_newuser_ios'
ws_fillna_newuser.title = 'fillna_and_sort_by_rank_ios'

for day in list_days_needed[:2]:
    print day

    # 得到该天iOS平台所有应用的新增量
    path_newuser_ios = '/Users/Alas/Documents/TD/iOS_Rank_Estimated_Model_V1/jupyter_sandbox_data/' +\
            'daily_data_from_interface/' + day[:7] + '/'
    path_newuser_ios_oneday = path_newuser_ios + day + '_newuser_iOS.txt'
    df_newuser_ios_oneday = pd.read_csv(path_newuser_ios_oneday, index_col=0)['newuser']
    df_newuser_ios_oneday.fillna(0, inplace=True)

    # 得到该天"待分析应用"的畅销排名
    df_rank_oneday = df_rank_needed[day]
    # 最终的数据类型为 DataFrame.Series, 方便以".ix"提取
    #print df_newuser_ios_oneday.head()
    #print df_rank_oneday.head()

    # 根据有排名的APP的ProductID提取"待分析应用"新增量
    list_newuser_ios_needed = []
    for appid in df_rank_oneday.index:
        try:
            list_newuser_ios_needed.append(df_newuser_ios_oneday.ix[appid])
        except Exception, e:
            print e, "can't find the appid in the file: " + day + '_newuser_iOS.txt'
            list_newuser_ios_needed.append(0)
    #print list_newuser_ios_needed

    # 比较待分析应用的排名列表和新增量列表
    list_tuple_of_rank_and_newuser_ios = zip(list(df_rank_oneday.index),\
            list(df_rank_oneday.values), list_newuser_ios_needed)
    #print len(list_tuple_of_rank_and_newuser_ios)
    #print list_tuple_of_rank_and_newuser_ios[:10]

    # 合并新增量跟排名值两列数据到一个DataFrame
    df_rank_newuser = pd.DataFrame(list_tuple_of_rank_and_newuser_ios,\
            columns=['productid', 'rank', 'newuser_ios',],)
    # 将productid 列设为行索引
    df_rank_newuser.index = df_rank_newuser['productid']
    df_rank_newuser = df_rank_newuser[['rank', 'newuser_ios']]

    # 按照新增量降序排
    df_rank_newuser.sort_values(by='newuser_ios', ascending=False, inplace=True)
#    print df_rank_newuser
    # 估计排名缺失的部分
    df_sort_by_newuser = FillnaRank(df_rank_newuser)
#    print df_sort_by_newuser

    # 估计新增量缺失的部分
    df_sort_by_rank = FillnaNewuser(df_sort_by_newuser)
#    print df_sort_by_rank

    # 调整异常值
    df_rank_newuser_revised = FilterOutliers(df_sort_by_rank)

    # 拟合
    FittingRevised(df_rank_newuser_revised)

"""
    # 将填充缺失值前的数据写入"sort_by_newuser_ios"工作表中
    for i in df_rank_newuser.index:
        list_to_write = list(df_rank_newuser.ix[i].values)
        list_to_write.insert(0, day)
        ws.append(list_to_write)
    # 将填充缺失值后的数据写入"fillna_rank_and_sort_by_newuser_ios"工作表中
    for i in df_sort_by_newuser.index:
        list_to_write = list(df_sort_by_newuser.ix[i].values)
        try:
            list_to_write[0] = round(list_to_write[0])
        except Exception, e:
            print e, "四舍五入失败"
        list_to_write.insert(0, day)
        ws_fillna_rank.append(list_to_write) 
    # 将填充缺失值前的数据写入"sort_by_rank"工作表中
    for i in df_sort_by_rank.index:
        list_to_write = list(df_sort_by_rank.ix[i].values)
        list_to_write.insert(0, day)
        ws_fillna_newuser.append(list_to_write)

wb.save(path_to_write + 'df_sort_by_newuser_revised_alldays.xlsx')
"""




